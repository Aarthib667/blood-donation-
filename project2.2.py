from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import os
from openpyxl import Workbook, load_workbook

# Function to calculate grade
def calculate_grade(avg):
    if avg >= 90:
        return 'A'
    elif avg >= 80:
        return 'B'
    elif avg >= 70:
        return 'C'
    elif avg >= 60:
        return 'D'
    else:
        return 'F'

# Function to check pass/fail
def check_pass(marks):
    return "Pass" if all(mark >= 35 for mark in marks) else "Fail"

# Save function
def save_data():
    name = entry_name.get().strip()
    roll_no = entry_roll.get().strip()

    try:
        marks = [int(entries[subject].get()) for subject in subjects]
    except ValueError:
        messagebox.showerror("Error", "Please enter valid numbers for marks.")
        return

    if not name or not roll_no:
        messagebox.showwarning("Input Error", "Name and Roll No are required!")
        return

    avg = sum(marks) / len(marks)
    grade = calculate_grade(avg)
    result = check_pass(marks)

    # Update Labels
    label_avg_val.config(text=f"{avg:.2f}")
    label_grade_val.config(text=grade)
    label_result_val.config(text=result)

    file_name = "student_records.xlsx"

    # Create or load workbook
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Records"
        ws.append(["Name", "Roll No"] + subjects + ["Average", "Grade", "Result"])
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    ws.append([name, roll_no] + marks + [avg, grade, result])
    wb.save(file_name)

    messagebox.showinfo("Success", "Student data saved successfully!")

# Tkinter setup
window = Tk()
window.title("Student Marksheet")
window.geometry("600x500")

# Notebook tabs
my_notebook = ttk.Notebook(window)
my_notebook.pack(expand=1, fill=BOTH)

# Student Marksheet tab
window1 = ttk.Frame(my_notebook)
my_notebook.add(window1, text="Student Marksheet")

# Subjects list
subjects = ["Maths", "Science", "English", "History", "Computer"]
entries = {}

# Form Fields
Label(window1, text="Name:").grid(row=0, column=0, padx=10, pady=5, sticky=W)
entry_name = Entry(window1)
entry_name.grid(row=0, column=1, pady=5)

Label(window1, text="Roll No:").grid(row=1, column=0, padx=10, pady=5, sticky=W)
entry_roll = Entry(window1)
entry_roll.grid(row=1, column=1, pady=5)

row_index = 2
for subject in subjects:
    Label(window1, text=f"{subject}:").grid(row=row_index, column=0, padx=10, pady=5, sticky=W)
    e = Entry(window1)
    e.grid(row=row_index, column=1, pady=5)
    entries[subject] = e
    row_index += 1

# Average, Grade, Result labels
Label(window1, text="Average:").grid(row=row_index, column=0, padx=10, pady=5, sticky=W)
label_avg_val = Label(window1, text="")
label_avg_val.grid(row=row_index, column=1, pady=5)
row_index += 1

Label(window1, text="Grade:").grid(row=row_index, column=0, padx=10, pady=5, sticky=W)
label_grade_val = Label(window1, text="")
label_grade_val.grid(row=row_index, column=1, pady=5)
row_index += 1

Label(window1, text="Result:").grid(row=row_index, column=0, padx=10, pady=5, sticky=W)
label_result_val = Label(window1, text="")
label_result_val.grid(row=row_index, column=1, pady=5)
row_index += 1

# Save button
Button(window1, text="Save Data", command=save_data).grid(row=row_index, column=0, columnspan=2, pady=15)

# Other tabs
window2 = ttk.Frame(my_notebook)
my_notebook.add(window2, text="Student Form")
window3 = ttk.Frame(my_notebook)
my_notebook.add(window3, text="Teachers Form")

window.mainloop()


